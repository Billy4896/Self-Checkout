import sqlite3
from Product import Product
from Transcation import Transcation
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt

class SupermarketDAO:
    """A SupermarketDAO class required for sqlite functionality"""

    def __init__(self):
        """Initialise attributes for supermarket class"""
        self.db = ""

    def setup_db(self):
        """Create the checkout_db database"""
        try:
            self.db = sqlite3.connect('checkout_db.db')
            cursor = self.db.cursor()

            # Creating the new product table within the checkout_db.
            cursor.execute('''CREATE TABLE IF NOT EXISTS 
                              product(barcode CHAR(20), productName CHAR(25) NOT NULL, unitPrice FLOAT(5,2))''')
            # Creating the new transaction table within the checkout_db.
            cursor.execute('''CREATE TABLE IF NOT EXISTS
                                transactions(barcode CHAR(20), date DATE, amount FLOAT(5,2))''')

            # List of dummy products to fill the product table
            product = [('123', 'Milk 2 Litres', 3.50),
                       ('456', 'Bread 500g', 2.50),
                       ('122', 'SoyMilk 1 Litre', 3.30),
                       ('100', 'Eggs 700g', 4.90),
                       ('101', 'Bacon 200g', 5.00),
                       ('102', 'Butter 500g', 4.75),
                       ('333', 'Hash Browns', 3.80),
                       ('808', 'Vanilla icecream 1 Litre', 12.00),
                       ('901', 'Toliet paper 24 pack', 15.50),
                       ('950', 'Cat litter', 15.00)]

            # Inserting the product information into the product table.
            cursor.executemany('''INSERT OR IGNORE INTO product(barcode, productName, unitPrice) VALUES (?,?,?)''',
                               product)
            self.db.commit()
        # Except block to catch any errors that may occur.
        except Exception as e:
            print(f"Unable to create the table requested. {e}")
            self.db.rollback()
        # Close the database.
        finally:
            self.db.close()

    def add_product_to_db(self, product):
        """Adds products to the product table in the checkout_db.db"""
        # Open the database with a try block and using an insert query, insert the product object data into the product table.
        try:
            self.db = sqlite3.connect('checkout_db.db')
            cursor = self.db.cursor()
            cursor.execute('''INSERT INTO product(barcode, productName, unitPrice) VALUES(?,?,?)''',
                           (product.get_barcode(), product.get_product_name(), product.get_price()))
            self.db.commit()
            print("Data insert was successful.")
        # Except block to display any errors that may occur.
        except Exception as e:
            print(f"An error has occurred. {e}")
        # Close the database.
        finally:
            self.db.close()

    def list_all_products(self):
        """Lists all the products in the product table of the checkout_db.db"""
        # Open the database with a try block and using a select query, select all product data ordered by barcode.
        product_list = []
        try:
            self.db = sqlite3.connect('checkout_db.db')
            cursor = self.db.cursor()
            cursor.execute('''SELECT * FROM product''')
            rows = cursor.fetchall()
            # If no rows can be found return nothing, else return a list of products.
            if rows == None:
                return None
            else:
                for row in rows:
                    product_list.append(Product(row[0], row[1], str(row[2])))
        # Except block to display any errors that may occur.
        except Exception as e:
            print(f"An error has occurred. {e}")
        # Close the database.
        finally:
            self.db.close()
            # Call the bubble sort algorithm to sort the list in ascending order based on barcode.
            products = self.bubble_sort_list_all_products(product_list)
            return products

    def bubble_sort_list_all_products(self, products):
        # Gets the length of the list and assign it to the variable n.
        n = len(products)
        # Begins the outer for loop that runs the algorithm (n – 1) times.
        for i in range(n - 1):
            # Once the list has been sorted flag will break the inner loop
            flag = 0
            # The inner loop compares all the values in the list.
            for j in range(n - 1):
                # If the value on the left-hand side is greater than the one on the immediate right side.
                if products[j].get_barcode() > products[j + 1].get_barcode():
                    # Assign the value of products[j] to a temporary variable
                    temp = products[j]
                    # Assign products[j + 1] to products[j]
                    products[j] = products[j + 1]
                    # Assign temp to products[j + 1]
                    products[j + 1] = temp
                    # Increase flag to 1 to indicate the swap was successful.
                    flag = 1
            # If the variable flag is 0 break the inner loop.
            if flag == 0:
                break
        # Return the sorted list.
        return products

    def find_product(self, barcode):
        """Finds products in the product table in the checkout_db.db"""
        # Convert the barcode string into a list.
        barcode = [barcode]
        # Open the database with a try block and using a select query return the barcode, productName and unitPrice from the product table using the [barcode] data.
        try:
            self.db = sqlite3.connect('checkout_db.db')
            cursor = self.db.cursor()
            cursor.execute('''SELECT barcode, productName, unitPrice FROM product WHERE barcode = ?''', barcode)
            rows = cursor.fetchone()
            # If no rows can be found return nothing, else return a products.
            if rows == None:
                return None
            else:
                rows2 = list(rows)
                product = Product(rows2[0], rows2[1], str(rows2[2]))
                return product
        # Except block to display any errors that may occur.
        except Exception as e:
            print(f"An error has occurred. {e}")
        # Close the database.
        finally:
            self.db.close()

    def list_all_transcations(self):
        """Lists all transactions in the transactions table in the checkout_db.db"""
        # Open the database with a try block and using a select query, select all transaction data ordered by date.
        transactions_list = []
        try:
            self.db = sqlite3.connect('checkout_db.db')
            cursor = self.db.cursor()
            cursor.execute('''SELECT date, barcode, amount FROM transactions''')
            rows = cursor.fetchall()
            # If no rows can be found return nothing, else return a list of transactions.
            if rows == None:
                return None
            else:
                for row in rows:
                    transactions_list.append(Transcation(row[0], row[1], str(row[2])))
        # Except block to display any errors that may occur.
        except Exception as e:
            print(f"An error has occurred. {e}")
        # Close the database.
        finally:
            self.db.close()
            # Call the selection sort algorithm to sort the list in ascending order based on date.
            transactions = self.selection_sort_list_all_transactions(transactions_list)
            return transactions

    def selection_sort_list_all_transactions(self, transactions):
        """Sorts the list all transactions function using a Selection Sort algorihm"""
        # Get the total number of elements in the list.
        n = len(transactions)
        # For every element in the indexing length, set the minimum to the i position.
        for i in range(n - 1):
            minValueIndex = i
            # For all the elements to the right-hand side of i, to the length of the list.
            for j in range(i + 1, len(transactions)):
                # If transaction object j is less than the current minimum value.
                if transactions[j].get_current_date() < transactions[minValueIndex].get_current_date():
                    # Replace the value of minValueIndex with transaction object j.
                    minValueIndex = j
            # If minValueIndex is no longer equal to i.
            if minValueIndex != i :
                # Assign the leftmost value in a temporary variable.
                temp = transactions[i].get_current_date()
                # The lower value from the right-hand side takes the i position.
                transactions[i] = transactions[minValueIndex]
                # The minimum value stored held in the temp variable is now stored in the position previously held by the minimum value.
                transactions[minValueIndex] = temp
        # Return a sorted list.
        return transactions

    def add_transactions_to_db(self, transaction):
        """Adds transcations to the transcations table in the checkout_db.db"""
        # Open the database with a try block and using an insert query, insert the transaction object data into the transaction table.
        try:
            self.db = sqlite3.connect('checkout_db.db')
            cursor = self.db.cursor()
            cursor.execute('''INSERT INTO transactions(date, barcode, amount) VALUES(?,?,?)''',
                           (transaction.get_current_date(), transaction.get_barcode(), transaction.get_amount()))
            self.db.commit()
        # Except block to display any errors that may occur.
        except Exception as e:
            print(f"An error has occurred. {e}")
            self.db.rollback()
        # Close the database.
        finally:
            self.db.close()

    def display_barchart_of_products_sold(self):
        """Creates and displays a barcart using Product data"""
        # Import data into a list through calling the list all transactions method.
        transaction = self.list_all_transcations()
        # Create the dictionary & define x & y-axis lists
        dict_of_transactions = {}
        x = []
        y = []
        # For each transaction record in the transaction list, get the barcode and assign it to the barcode variable to be used as the dictionary key.
        for trans in transaction:
            barcode = trans.get_barcode()
            # If the barcode is not in the dictionary, place it inside the dictionary as the key and assign its value as 1.
            if barcode not in dict_of_transactions:
                dict_of_transactions[barcode] = 1
            # Else if the barcode already exists increase its value by 1.
            else:
                dict_of_transactions[barcode] = dict_of_transactions[barcode] + 1
        # For loop to append each key and its value to the x & y-axis.
        for barcode in dict_of_transactions:
            x.append(self.find_product(barcode).get_product_name())
            y.append(dict_of_transactions[barcode])
        # Plot the chart and show it.
        plt.bar(x, y)
        plt.show()

    def display_excel_report_of_transactions(self):
        """Displays excel report to transactions"""
        # Import a new instance of workbook overwriting the previous to prevent the same data being recorded twice.
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Date"
        sheet["B1"] = "Barcode"
        sheet["C1"] = "Amount"
        # Save the workbook.
        workbook.save('ReportOfTransactions.xlsx')
        # Import an instance of Supermarket and run the def list_all_transactions to a new list.
        transactions_list = []
        rows = self.list_all_transcations()
        for row in rows:
            x = row.get_current_date(), row.get_barcode(), row.get_amount()
            transactions_list.append(x)
        # Load the workbook used to display the transactions.
        workbook = load_workbook(filename="ReportOfTransactions.xlsx")
        # Activate the workbook.
        sheet = workbook.active
        # Append the transactions in the rows list.
        for product in transactions_list:
            sheet.append(product)
        # Output the workbook to python.
        for value in sheet.iter_rows(min_row=1,
                                     min_col=1,
                                     max_col=3,
                                     values_only=True):
            print(value)
        # Save the workbook.
        workbook.save('ReportOfTransactions.xlsx')

